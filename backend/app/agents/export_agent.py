from __future__ import annotations

import io
from datetime import datetime
from typing import Iterable

from flask import abort
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from ..agents.responses import agent_response
from ..extensions import db
from ..models import Material, PublicityBatch, User
from ..state_machine import MaterialStatus


CATEGORY_ORDER = ["德育", "智育", "体育", "美育", "劳育", "能力"]
CATEGORY_CAPS = {
    "德育": 15.0,
    "智育": 20.0,
    "体育": 8.0,
    "美育": 6.0,
    "劳育": 6.0,
    "能力": 8.0,
}


def _category_breakdown(materials: list[Material]) -> dict[str, float]:
    totals = {category: 0.0 for category in CATEGORY_ORDER}
    for material in materials:
        if material.status in {MaterialStatus.APPROVED.value, MaterialStatus.PUBLICIZING.value, MaterialStatus.PUBLICITY_ENDED.value}:
            totals[material.category] = totals.get(material.category, 0.0) + float(material.score or 0)
    return totals


def _total_score(materials: list[Material]) -> float:
    return round(sum(_category_breakdown(materials).values()), 2)


def _resolve_user(user: User) -> User:
    return user


def _student_export_payload(user: User, term_id: int | None) -> dict:
    query = Material.query.filter_by(student_id=user.id)
    if term_id:
        query = query.filter_by(term_id=term_id)
    materials = query.order_by(Material.created_at.asc()).all()
    breakdown = _category_breakdown(materials)
    return {
        "student": user.to_dict(),
        "materials": [m.to_dict(include_student=False) for m in materials],
        "breakdown": breakdown,
        "breakdownWithCap": [
            {"category": c, "score": round(breakdown.get(c, 0.0), 2), "cap": CATEGORY_CAPS[c]}
            for c in CATEGORY_ORDER
        ],
        "totalScore": _total_score(materials),
        "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def _ranking_payload(user: User | None, term_id: int | None, anonymous: bool) -> dict:
    query = (
        db.session.query(User, db.func.coalesce(db.func.sum(Material.score), 0).label("total"))
        .join(Material, Material.student_id == User.id)
        .filter(
            Material.status.in_(
                [
                    MaterialStatus.APPROVED.value,
                    MaterialStatus.PUBLICIZING.value,
                    MaterialStatus.PUBLICITY_ENDED.value,
                ]
            )
        )
    )
    if term_id:
        query = query.filter(Material.term_id == term_id)
    if user and user.role in {"student", "counselor"} and user.class_group_id:
        query = query.filter(User.class_group_id == user.class_group_id)
    rows = query.group_by(User.id).order_by(db.func.coalesce(db.func.sum(Material.score), 0).desc(), User.student_no.asc()).all()
    items = []
    for index, (student, total) in enumerate(rows, start=1):
        if anonymous:
            name = f"{student.name[:1]}*"
            student_no = student.student_no
        else:
            name = student.name
            student_no = student.student_no
        items.append(
            {
                "rank": index,
                "name": name,
                "studentNo": student_no,
                "className": student.class_name,
                "totalScore": float(total),
            }
        )
    batch = PublicityBatch.query.order_by(PublicityBatch.starts_at.desc()).first()
    return {
        "items": items,
        "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "batchTitle": batch.title if batch else "综测公示",
        "anonymous": anonymous,
    }


class ExportAgent:
    def student_summary(self, user: User, term_id: int | None = None) -> dict:
        return _student_export_payload(user, term_id)

    def student_pdf(self, user: User, term_id: int | None = None) -> bytes:
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        payload = _student_export_payload(user, term_id)
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=18 * mm,
            rightMargin=18 * mm,
            topMargin=18 * mm,
            bottomMargin=18 * mm,
            title=f"综测成绩单-{user.student_no}",
        )
        styles = getSampleStyleSheet()
        body_style = ParagraphStyle("body", parent=styles["BodyText"], fontName="STSong-Light", fontSize=10, leading=14)
        title_style = ParagraphStyle("title", parent=styles["Title"], fontName="STSong-Light", fontSize=18, leading=22, alignment=1)
        h2_style = ParagraphStyle("h2", parent=styles["Heading2"], fontName="STSong-Light", fontSize=13, leading=18, spaceBefore=8, spaceAfter=6)

        story = []
        story.append(Paragraph("高校综合测评成绩单", title_style))
        story.append(Paragraph(f"姓名：{user.name}　　学号：{user.student_no}　　班级：{user.class_name or '-'}", body_style))
        story.append(Paragraph(f"生成时间：{payload['generatedAt']}", body_style))
        story.append(Spacer(1, 6 * mm))

        story.append(Paragraph("五育得分汇总", h2_style))
        summary_data = [["五育类别", "得分", "上限"]]
        for item in payload["breakdownWithCap"]:
            summary_data.append([item["category"], f"{item['score']:.2f}", f"{item['cap']:.2f}"])
        summary_data.append(["合计", f"{payload['totalScore']:.2f}", "/"])
        summary_table = Table(summary_data, colWidths=[40 * mm, 30 * mm, 30 * mm])
        summary_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e9efff")),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#f3f6fc")),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#dde3ed")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dde3ed")),
                ]
            )
        )
        story.append(summary_table)
        story.append(Spacer(1, 6 * mm))

        story.append(Paragraph("材料明细", h2_style))
        detail_header = ["材料", "类别", "级别", "角色", "建议分", "状态", "证书编号", "发证日期"]
        detail_rows = [detail_header]
        for material in payload["materials"]:
            detail_rows.append(
                [
                    material["title"],
                    material["category"],
                    material.get("category", ""),
                    "-",
                    f"{float(material['score']):.2f}",
                    material["status"],
                    material["certificateNo"] or "-",
                    material["issuedAt"],
                ]
            )
        if len(detail_rows) == 1:
            detail_rows.append(["-", "-", "-", "-", "-", "-", "-", "-"])
        detail_table = Table(detail_rows, repeatRows=1, colWidths=[42 * mm, 18 * mm, 18 * mm, 18 * mm, 18 * mm, 22 * mm, 30 * mm, 24 * mm])
        detail_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#172033")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#dde3ed")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dde3ed")),
                ]
            )
        )
        story.append(detail_table)
        story.append(Spacer(1, 10 * mm))
        story.append(Paragraph("辅导员签字：____________________", body_style))
        story.append(Paragraph("日期：____________________", body_style))

        doc.build(story)
        buffer.seek(0)
        return buffer.read()

    def student_xlsx(self, user: User, term_id: int | None = None) -> bytes:
        payload = _student_export_payload(user, term_id)
        wb = Workbook()
        ws = wb.active
        ws.title = "综测成绩单"
        ws["A1"] = f"{user.name}（{user.student_no}）综测成绩单"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:E1")
        ws["A2"] = f"班级：{user.class_name or '-'}　生成时间：{payload['generatedAt']}"

        headers = ["五育类别", "得分", "上限"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E9EFFF")
            cell.alignment = Alignment(horizontal="center")
        for index, item in enumerate(payload["breakdownWithCap"], start=5):
            ws.cell(row=index, column=1, value=item["category"])
            ws.cell(row=index, column=2, value=float(item["score"]))
            ws.cell(row=index, column=3, value=float(item["cap"]))
        total_row = 4 + len(payload["breakdownWithCap"]) + 1
        ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
        ws.cell(row=total_row, column=2, value=float(payload["totalScore"])).font = Font(bold=True)
        ws.cell(row=total_row, column=3, value="/")

        detail_start = total_row + 3
        ws.cell(row=detail_start, column=1, value="材料明细").font = Font(bold=True, size=12)
        detail_headers = ["材料", "五育", "级别", "角色", "建议分", "状态", "证书编号", "发证日期", "备注"]
        for col, header in enumerate(detail_headers, start=1):
            cell = ws.cell(row=detail_start + 1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E9EFFF")
            cell.alignment = Alignment(horizontal="center")
        for offset, material in enumerate(payload["materials"], start=2):
            row = detail_start + offset
            ws.cell(row=row, column=1, value=material["title"])
            ws.cell(row=row, column=2, value=material["category"])
            ws.cell(row=row, column=3, value=material.get("category", ""))
            ws.cell(row=row, column=4, value="-")
            ws.cell(row=row, column=5, value=float(material["score"]))
            ws.cell(row=row, column=6, value=material["status"])
            ws.cell(row=row, column=7, value=material["certificateNo"] or "-")
            ws.cell(row=row, column=8, value=material["issuedAt"])
            ws.cell(row=row, column=9, value=(material.get("description") or "")[:60])

        for col in range(1, len(detail_headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["G"].width = 22

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def ranking_pdf(self, user: User | None, term_id: int | None, anonymous: bool) -> bytes:
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        payload = _ranking_payload(user, term_id, anonymous)
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=18 * mm,
            rightMargin=18 * mm,
            topMargin=18 * mm,
            bottomMargin=18 * mm,
            title="综测排行榜",
        )
        styles = getSampleStyleSheet()
        body_style = ParagraphStyle("body", parent=styles["BodyText"], fontName="STSong-Light", fontSize=10, leading=14)
        title_style = ParagraphStyle("title", parent=styles["Title"], fontName="STSong-Light", fontSize=18, leading=22, alignment=1)

        story = [Paragraph(payload["batchTitle"] or "综测排行榜", title_style), Spacer(1, 4 * mm)]
        story.append(Paragraph(f"生成时间：{payload['generatedAt']}　匿名：{'是' if anonymous else '否'}", body_style))
        story.append(Spacer(1, 6 * mm))

        rows = [["排名", "学号", "姓名", "班级", "总分"]]
        for item in payload["items"]:
            rows.append([str(item["rank"]), item["studentNo"], item["name"], item["className"] or "-", f"{item['totalScore']:.2f}"])
        table = Table(rows, repeatRows=1, colWidths=[20 * mm, 38 * mm, 30 * mm, 30 * mm, 30 * mm])
        table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#172033")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#dde3ed")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dde3ed")),
                ]
            )
        )
        story.append(table)
        doc.build(story)
        buffer.seek(0)
        return buffer.read()

    def ranking_xlsx(self, user: User | None, term_id: int | None, anonymous: bool) -> bytes:
        payload = _ranking_payload(user, term_id, anonymous)
        wb = Workbook()
        ws = wb.active
        ws.title = "综测排行榜"
        ws["A1"] = payload["batchTitle"] or "综测排行榜"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:E1")
        ws["A2"] = f"生成时间：{payload['generatedAt']}　匿名：{'是' if anonymous else '否'}"

        headers = ["排名", "学号", "姓名", "班级", "总分"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E9EFFF")
            cell.alignment = Alignment(horizontal="center")
        for offset, item in enumerate(payload["items"], start=5):
            ws.cell(row=offset, column=1, value=item["rank"])
            ws.cell(row=offset, column=2, value=item["studentNo"])
            ws.cell(row=offset, column=3, value=item["name"])
            ws.cell(row=offset, column=4, value=item["className"] or "-")
            ws.cell(row=offset, column=5, value=float(item["totalScore"]))

        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["D"].width = 24

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()
